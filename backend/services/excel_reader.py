"""
Reads uploaded Excel workbooks and converts them into the dict structures
expected by the audit engine.

Column-header matching is case-insensitive and tolerant of minor naming
variations.  Each ``parse_*`` function returns the same shape the audit
engine already consumes from JSON.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ── helpers ──────────────────────────────────────────────────────────────

def _norm(header: str) -> str:
    """Lowercase, strip, collapse whitespace, drop non-alnum except underscore."""
    return re.sub(r"[^a-z0-9_]", "", header.strip().lower().replace(" ", "_"))


def _read_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Return a list of row-dicts keyed by normalised header names."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [_norm(str(h or "")) for h in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def _date_str(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def _float(val: Any, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int(val: Any, default: int = 0) -> int:
    return int(_float(val, float(default)))


def _bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("yes", "true", "1", "y")
    return bool(val)


def _str(val: Any, default: str = "") -> str:
    if val is None:
        return default
    return str(val).strip()


# ── sheet-name helpers ───────────────────────────────────────────────────

def _find_sheet(wb, *candidates: str) -> Worksheet | None:
    """Return the first sheet whose normalised name matches a candidate."""
    norm_map = {_norm(name): name for name in wb.sheetnames}
    for c in candidates:
        real = norm_map.get(_norm(c))
        if real:
            return wb[real]
    return None


# ── public parsers ───────────────────────────────────────────────────────


def parse_quickbooks(path: str | Path) -> dict:
    """
    Parse the QuickBooks revenue workbook.
    Expected sheet: "Revenue Transactions"
    Returns the same shape as quickbooks_revenue.json.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _find_sheet(wb, "Revenue Transactions", "Revenue", "Transactions")
    if ws is None:
        ws = wb.worksheets[0]

    rows = _read_sheet(ws)
    wb.close()

    transactions = []
    for r in rows:
        transactions.append({
            "id": _str(r.get("transaction_id", r.get("txn_id", r.get("id")))),
            "date": _date_str(r.get("date")),
            "customer": _str(r.get("customer")),
            "amount": _float(r.get("amount")),
            "channel": _str(r.get("channel", "")).lower(),
            "description": _str(r.get("description", "")),
        })

    return {
        "company": "FreshGlow Skincare",
        "period": _guess_period(transactions),
        "transactions": transactions,
    }


def parse_sales_channels(path: str | Path) -> tuple[dict, dict]:
    """
    Parse the sales channel invoices workbook (Shopify + Amazon sheets).
    Returns (shopify_dict, amazon_dict) in the shapes of
    shopify_invoices.json and amazon_orders.json.
    """
    wb = load_workbook(path, data_only=True, read_only=True)

    shopify_ws = _find_sheet(wb, "Shopify Orders", "Shopify", "Shopify Invoices")
    amazon_ws = _find_sheet(wb, "Amazon Orders", "Amazon", "Amazon Invoices")

    shopify_invoices: list[dict] = []
    if shopify_ws:
        for r in _read_sheet(shopify_ws):
            inv: dict[str, Any] = {
                "order_id": _str(r.get("order_id")),
                "txn_ref": _str(r.get("qb_reference", r.get("txn_ref", r.get("qbreference")))),
                "date": _date_str(r.get("date")),
                "customer": _str(r.get("customer")),
                "amount": _float(r.get("final_amount", r.get("amount"))),
                "status": _str(r.get("status", "")).lower(),
            }
            discount = _float(r.get("discount", r.get("discount_applied", 0)))
            if discount:
                inv["discount_applied"] = discount
            shopify_invoices.append(inv)

    amazon_orders: list[dict] = []
    if amazon_ws:
        for r in _read_sheet(amazon_ws):
            order: dict[str, Any] = {
                "order_id": _str(r.get("order_id")),
                "txn_ref": _str(r.get("qb_reference", r.get("txn_ref", r.get("qbreference")))),
                "date": _date_str(r.get("order_date", r.get("date"))),
                "customer": _str(r.get("customer")),
                "amount": _float(r.get("amount")),
                "status": _str(r.get("status", "")).lower(),
            }
            booked = _date_str(r.get("date_booked_in_qb", r.get("date_booked")))
            if booked:
                order["date_booked"] = booked
            amazon_orders.append(order)

    wb.close()
    return {"invoices": shopify_invoices}, {"orders": amazon_orders}


def parse_inventory(path: str | Path) -> dict:
    """
    Parse the inventory workbook (Inventory Snapshot + Purchase Orders).
    Returns the same shape as inventory_records.json.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    snap_ws = _find_sheet(wb, "Inventory Snapshot", "Inventory", "Snapshot")
    po_ws = _find_sheet(wb, "Purchase Orders", "POs", "PO")

    items: list[dict] = []
    if snap_ws:
        for r in _read_sheet(snap_ws):
            items.append({
                "sku": _str(r.get("sku")),
                "name": _str(r.get("product_name", r.get("name"))),
                "category": _str(r.get("abc_category", r.get("category", "C"))).upper(),
                "system_count": _int(r.get("system_count")),
                "expected_count": _int(r.get("expected_count")),
                "unit_cost": _float(r.get("unit_cost")),
            })

    purchase_orders: list[dict] = []
    if po_ws:
        for r in _read_sheet(po_ws):
            purchase_orders.append({
                "po_id": _str(r.get("po_number", r.get("po_id"))),
                "supplier": _str(r.get("supplier")),
                "sku": _str(r.get("sku")),
                "ordered_qty": _int(r.get("qty_ordered", r.get("ordered_qty"))),
                "received_qty": _int(r.get("qty_received", r.get("received_qty"))),
                "invoice_amount": _float(r.get("invoice_amount")),
                "date_ordered": _date_str(r.get("order_date", r.get("date_ordered"))),
                "date_received": _date_str(r.get("received_date", r.get("date_received"))),
            })

    wb.close()
    return {
        "warehouse": "ShipBob - LA",
        "snapshot_date": _guess_snapshot_date(items),
        "items": items,
        "purchase_orders": purchase_orders,
    }


def parse_payroll(path: str | Path) -> dict:
    """
    Parse the payroll workbook (Employee Roster + Payroll + GL Reconciliation).
    Returns the same shape as payroll_records.json.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    roster_ws = _find_sheet(wb, "Employee Roster", "Roster", "Employees")
    payroll_ws = _find_sheet(wb, "March 2026 Payroll", "Payroll", "Pay")
    gl_ws = _find_sheet(wb, "GL Reconciliation", "GL", "General Ledger")

    roster_rows = _read_sheet(roster_ws) if roster_ws else []
    payroll_rows = _read_sheet(payroll_ws) if payroll_ws else []
    gl_rows = _read_sheet(gl_ws) if gl_ws else []

    wb.close()

    roster_map = {_str(r.get("employee_id", r.get("id"))): r for r in roster_rows}
    payroll_map = {_str(r.get("employee_id", r.get("id"))): r for r in payroll_rows}

    all_ids = list(dict.fromkeys(list(roster_map.keys()) + list(payroll_map.keys())))

    employees: list[dict] = []
    for eid in all_ids:
        ros = roster_map.get(eid, {})
        pay = payroll_map.get(eid, {})
        emp: dict[str, Any] = {
            "id": eid,
            "name": _str(ros.get("name", pay.get("name"))),
            "type": _str(ros.get("type", pay.get("type", "W2"))),
            "title": _str(ros.get("title", "")),
            "status": _str(ros.get("status", "active")).lower(),
            "gross_pay": _float(pay.get("gross_pay", 0)),
            "federal_tax": _float(pay.get("federal_tax", 0)),
            "state_tax": _float(pay.get("state_tax_ca", pay.get("state_tax", 0))),
            "start_date": _date_str(ros.get("start_date")),
        }

        term = _date_str(ros.get("termination_date"))
        if term:
            emp["termination_date"] = term

        hours = _float(ros.get("hours_week", ros.get("hoursweek", 0)))
        if hours:
            emp["hours_per_week"] = hours

        equip = ros.get("uses_company_equipment", ros.get("usescompanyequipment"))
        if equip is not None:
            emp["uses_company_equipment"] = _bool(equip)

        sdi_rate = _float(pay.get("sdi_rate_applied", 0))
        correct_sdi = _float(pay.get("correct_sdi_rate", 0))
        if sdi_rate:
            emp["ca_sdi_rate"] = sdi_rate
        if correct_sdi:
            emp["ca_sdi_correct_rate"] = correct_sdi

        employees.append(emp)

    payroll_total = sum(e["gross_pay"] for e in employees)
    gl_payroll_entry = payroll_total

    for row in gl_rows:
        source = _str(row.get("source", "")).lower()
        if "payroll" in source:
            payroll_total = _float(row.get("amount", payroll_total))
        elif "gl" in source or "ledger" in source or "general" in _str(row.get("account", "")).lower():
            gl_payroll_entry = _float(row.get("amount", gl_payroll_entry))

    if gl_rows and len(gl_rows) >= 2:
        amounts = [_float(r.get("amount", 0)) for r in gl_rows if _float(r.get("amount", 0)) > 0]
        if len(amounts) >= 2:
            amounts_sorted = sorted(amounts, reverse=True)
            payroll_total = amounts_sorted[0]
            gl_payroll_entry = amounts_sorted[1]

    return {
        "pay_period": "March 2026",
        "employees": employees,
        "payroll_total": payroll_total,
        "gl_payroll_entry": gl_payroll_entry,
    }


# ── file classification ──────────────────────────────────────────────────

CLASSIFICATION_KEYWORDS = {
    "quickbooks": ["quickbooks", "revenue_transaction", "revenue transactions"],
    "sales_channels": ["shopify", "amazon", "sales_channel", "invoices"],
    "inventory": ["inventory", "purchase_order", "warehouse", "snapshot"],
    "payroll": ["payroll", "employee_roster", "gl_reconciliation", "roster"],
}


def classify_workbook(path: str | Path) -> str | None:
    """
    Guess the workbook type based on filename and sheet names.
    Returns one of: 'quickbooks', 'sales_channels', 'inventory', 'payroll', or None.
    """
    p = Path(path)
    filename_lower = p.stem.lower()
    wb = load_workbook(path, read_only=True)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    wb.close()

    combined = filename_lower + " " + " ".join(sheet_names_lower)

    for category, keywords in CLASSIFICATION_KEYWORDS.items():
        if any(kw in combined for kw in keywords):
            return category
    return None


def summarize_workbook(path: str | Path) -> dict:
    """Return a quick summary of sheets, row counts, and column headers."""
    wb = load_workbook(path, data_only=True, read_only=True)
    summary: dict[str, Any] = {"filename": Path(path).name, "sheets": []}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) for h in (rows[0] if rows else []) if h is not None]
        summary["sheets"].append({
            "name": name,
            "row_count": max(0, len(rows) - 1),
            "columns": headers,
        })
    wb.close()
    return summary


# ── convenience ──────────────────────────────────────────────────────────

def _guess_period(transactions: list[dict]) -> str:
    for t in transactions:
        d = t.get("date")
        if d:
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                return dt.strftime("%B %Y")
            except ValueError:
                continue
    return "Unknown"


def _guess_snapshot_date(items: list[dict]) -> str:
    return date.today().isoformat()
