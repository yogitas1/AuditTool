"""
OpenAI function-calling audit agent.

The agent works with Excel files at their original locations on the
filesystem.  Files can be registered by scanning a local directory
or by uploading through the web UI.  All edits are made in place.
"""

from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI

from .audit_engine import run_inventory_audit, run_payroll_audit, run_revenue_audit
from .excel_reader import (
    classify_workbook,
    parse_inventory,
    parse_payroll,
    parse_quickbooks,
    parse_sales_channels,
    summarize_workbook,
)
from .excel_writer import (
    edit_cell,
    get_corrections_log as _writer_get_log,
    clear_corrections_log as _writer_clear_log,
)

load_dotenv()

# ── config ───────────────────────────────────────────────────────────────

MODEL = "gpt-4o"
MAX_TOKENS = 2500
TEMPERATURE = 0.2
UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
try:
    UPLOAD_DIR.mkdir(exist_ok=True)
except (OSError, PermissionError):
    pass  # read-only filesystem (e.g. Vercel serverless)

# ── file registry ────────────────────────────────────────────────────────
# Maps filename -> absolute Path.  Populated by scan_directory() or
# register_upload().  The agent resolves filenames through this registry
# so it can work with files anywhere on the filesystem.

_file_registry: dict[str, Path] = {}


def scan_directory(directory: str) -> list[dict]:
    """Find all .xlsx files in a directory and register them."""
    d = Path(directory).expanduser().resolve()
    if not d.is_dir():
        raise FileNotFoundError(f"Not a directory: {d}")

    found: list[dict] = []
    for fp in sorted(d.glob("*.xlsx")):
        _file_registry[fp.name] = fp
        found.append(summarize_workbook(fp))
    return found


def register_upload(filepath: Path) -> None:
    """Register a file that was uploaded via the web UI."""
    _file_registry[filepath.name] = filepath.resolve()


def get_registered_files() -> list[dict]:
    """Return summaries of all registered files."""
    result = []
    for name, fp in sorted(_file_registry.items()):
        if fp.exists():
            s = summarize_workbook(fp)
            s["path"] = str(fp)
            result.append(s)
    return result


def has_files() -> bool:
    return bool(_file_registry)


# ── tool definitions (OpenAI function-calling schema) ────────────────────

TOOLS: list[dict] = [
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "List all Excel files currently registered for auditing, with their sheet names and row counts.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "inspect_file",
            "description": (
                "Inspect a registered Excel file — returns sheet names, "
                "column headers, and row counts."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Name of the file (e.g. 'payroll_report_march2026.xlsx')",
                    }
                },
                "required": ["filename"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "load_and_audit",
            "description": (
                "Parse the registered Excel files and run audit checks "
                "(revenue, inventory, payroll, or all). Returns findings."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "audit_type": {
                        "type": "string",
                        "enum": ["revenue", "inventory", "payroll", "all"],
                        "description": "Which audit to run.",
                    },
                },
                "required": ["audit_type"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_excel_data",
            "description": (
                "Read raw row data from a specific sheet inside a registered "
                "Excel file. Useful for answering detailed questions."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Name of the file.",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet to read (optional — reads first sheet if omitted).",
                    },
                },
                "required": ["filename"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "apply_correction",
            "description": (
                "Edit a single cell in a registered Excel file to fix a discrepancy. "
                "The change is saved to the original file immediately and logged."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "Name of the Excel file to edit.",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name to edit.",
                    },
                    "row_identifier_column": {
                        "type": "string",
                        "description": "Column used to locate the row (e.g. 'Transaction ID').",
                    },
                    "row_identifier_value": {
                        "type": "string",
                        "description": "Value to match in the identifier column (e.g. 'TXN-002').",
                    },
                    "target_column": {
                        "type": "string",
                        "description": "Column whose value should be changed.",
                    },
                    "new_value": {
                        "type": "string",
                        "description": "The corrected value to write.",
                    },
                    "reason": {
                        "type": "string",
                        "description": "Why this correction is being made.",
                    },
                },
                "required": [
                    "filename", "sheet_name", "row_identifier_column",
                    "row_identifier_value", "target_column", "new_value", "reason",
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "auto_fix_all",
            "description": (
                "Run the specified audit, then automatically apply corrections "
                "for every fixable finding directly in the original Excel files. "
                "Returns a summary of all corrections made."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "audit_type": {
                        "type": "string",
                        "enum": ["revenue", "inventory", "payroll", "all"],
                        "description": "Which audit to run and auto-fix.",
                    },
                },
                "required": ["audit_type"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_corrections_log",
            "description": "Return the full log of all corrections applied during this session.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
]

# ── module-level state ───────────────────────────────────────────────────

_client: OpenAI | None = None
conversation_history: list[dict[str, Any]] = []
_parsed_cache: dict[str, dict] = {}

SYSTEM_PROMPT = """\
You are **AuditAI**, an autonomous financial audit agent for small and mid-size businesses.

You have tools to read, analyze, and **directly fix** Excel files on the user's
local filesystem.  You operate **without requiring human approval** — you find
discrepancies, edit the original files **in place**, and report what you changed.

IMPORTANT: You edit the user's original files directly at their real locations
on disk. You do NOT create copies or new files.

## Your workflow

1. **List files** to see what spreadsheets are available.
2. **Inspect** files to understand their structure (sheets, columns, row counts).
3. **Run audits** — the engine cross-references revenue with invoices, inventory
   with purchase orders, and payroll with GL entries.
4. **Auto-fix** every correctable discrepancy directly in the original files using
   `apply_correction` (individual) or `auto_fix_all` (bulk).
5. **Report** a clear summary of all changes, grouped by audit area.
6. **Answer follow-ups** using the raw data or the corrections log.

## What you can fix autonomously

- **Revenue:** Amount mismatches, authorized discounts, wrong-period bookings.
- **Inventory:** System count discrepancies, PO receipt mismatches.
- **Payroll:** Terminated employee payments, incorrect SDI rates, GL reconciliation.

## What you should flag but NOT auto-fix

- Unmatched transactions with no source document.
- Contractor misclassification (legal/HR decision).
- Discrepancies you aren't confident about.

## Output format

- Use markdown. Be concise but thorough.
- Reference specific IDs (TXN-002, EMP-005, SKU HA-SERUM-50, etc.).
- After fixing, show a **Corrections Summary** table.
- Remind the user that the original files on their laptop have been updated.
"""

# ── internal helpers ─────────────────────────────────────────────────────

def _get_client() -> OpenAI:
    global _client
    if _client is None:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY is not set.")
        _client = OpenAI(api_key=api_key)
    return _client


def _all_files() -> list[Path]:
    """Return all registered file paths that still exist."""
    return [fp for fp in _file_registry.values() if fp.exists()]


def _resolve(filename: str) -> Path:
    """Look up a filename in the registry. Returns the absolute path."""
    fp = _file_registry.get(filename)
    if fp and fp.exists():
        return fp
    raise FileNotFoundError(
        f"File not found: {filename}. "
        f"Registered files: {list(_file_registry.keys())}"
    )


def _load_all_excel_data() -> dict[str, dict]:
    if _parsed_cache:
        return _parsed_cache

    for fp in _all_files():
        cat = classify_workbook(fp)
        if cat == "quickbooks":
            _parsed_cache["quickbooks"] = parse_quickbooks(fp)
        elif cat == "sales_channels":
            shopify, amazon = parse_sales_channels(fp)
            _parsed_cache["shopify"] = shopify
            _parsed_cache["amazon"] = amazon
        elif cat == "inventory":
            _parsed_cache["inventory"] = parse_inventory(fp)
        elif cat == "payroll":
            _parsed_cache["payroll"] = parse_payroll(fp)
    return _parsed_cache


def _run_audit(audit_type: str) -> list[dict]:
    data = _load_all_excel_data()

    qb = data.get("quickbooks")
    shopify = data.get("shopify")
    amazon = data.get("amazon")
    inv = data.get("inventory")
    pay = data.get("payroll")

    findings: list[dict] = []
    if audit_type in ("revenue", "all"):
        rev = run_revenue_audit(qb_data=qb, shopify_data=shopify, amazon_data=amazon)
        findings.extend({"_audit": "revenue", **f} for f in rev)
    if audit_type in ("inventory", "all"):
        inv_f = run_inventory_audit(inv_data=inv)
        findings.extend({"_audit": "inventory", **f} for f in inv_f)
    if audit_type in ("payroll", "all"):
        pay_f = run_payroll_audit(pay_data=pay)
        findings.extend({"_audit": "payroll", **f} for f in pay_f)
    return findings


# ── auto-fix logic ───────────────────────────────────────────────────────

def _find_file_for_category(category: str) -> Path | None:
    for fp in _all_files():
        cat = classify_workbook(fp)
        if cat == category:
            return fp
    return None


def _auto_fix_findings(audit_type: str) -> dict[str, Any]:
    invalidate_cache()
    findings = _run_audit(audit_type)
    applied: list[dict] = []
    skipped: list[dict] = []

    for f in findings:
        result = _try_fix_finding(f)
        if result:
            applied.append(result)
        else:
            skipped.append({
                "type": f.get("type"),
                "explanation": f.get("explanation", ""),
                "reason_skipped": "Cannot be auto-fixed — requires human investigation.",
            })

    invalidate_cache()

    return {
        "audit_type": audit_type,
        "total_findings": len(findings),
        "auto_fixed": len(applied),
        "skipped": len(skipped),
        "corrections": applied,
        "needs_review": skipped,
    }


def _try_fix_finding(f: dict) -> dict | None:
    ftype = f.get("type", "")
    audit = f.get("_audit", "")

    if ftype == "amount_mismatch" and audit == "revenue":
        qb_file = _find_file_for_category("quickbooks")
        if not qb_file:
            return None
        source_amount = f.get("source_amount")
        if source_amount is None:
            return None
        discount = 0
        if f.get("severity") == "medium":
            discount = abs(f.get("discrepancy", 0))
        corrected_amount = source_amount + discount if discount else source_amount
        result = edit_cell(
            filepath=qb_file,
            sheet_name="Revenue Transactions",
            row_identifier_column="Transaction ID",
            row_identifier_value=f["txn_id"],
            target_column="Amount",
            new_value=str(corrected_amount),
            reason=f"Corrected to match source invoice. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    if ftype == "wrong_period" and audit == "revenue":
        sales_file = _find_file_for_category("sales_channels")
        if not sales_file:
            return None
        correct_date = f.get("transaction_date")
        if not correct_date:
            return None
        result = edit_cell(
            filepath=sales_file,
            sheet_name="Amazon Orders",
            row_identifier_column="QB Reference",
            row_identifier_value=f["txn_id"],
            target_column="Date Booked in QB",
            new_value=correct_date,
            reason=f"Corrected booking date to match transaction date. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    if ftype in ("count_shortage", "count_overage") and audit == "inventory":
        inv_file = _find_file_for_category("inventory")
        if not inv_file:
            return None
        result = edit_cell(
            filepath=inv_file,
            sheet_name="Inventory Snapshot",
            row_identifier_column="SKU",
            row_identifier_value=f["sku"],
            target_column="System Count",
            new_value=str(f["expected_count"]),
            reason=f"Aligned system count with expected. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    if ftype == "terminated_employee_paid" and audit == "payroll":
        pay_file = _find_file_for_category("payroll")
        if not pay_file:
            return None
        result = edit_cell(
            filepath=pay_file,
            sheet_name="March 2026 Payroll",
            row_identifier_column="Employee ID",
            row_identifier_value=f["employee_id"],
            target_column="Gross Pay",
            new_value="0",
            reason=f"Stopped payment for terminated employee. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    if ftype == "incorrect_ca_sdi_rate" and audit == "payroll":
        pay_file = _find_file_for_category("payroll")
        if not pay_file:
            return None
        result = edit_cell(
            filepath=pay_file,
            sheet_name="March 2026 Payroll",
            row_identifier_column="Employee ID",
            row_identifier_value=f["employee_id"],
            target_column="SDI Rate Applied",
            new_value=str(f["correct_rate"]),
            reason=f"Corrected SDI rate from {f['actual_rate']} to {f['correct_rate']}. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    if ftype == "payroll_gl_discrepancy" and audit == "payroll":
        pay_file = _find_file_for_category("payroll")
        if not pay_file:
            return None
        corrected_total = f.get("gl_payroll_entry")
        if corrected_total is None:
            return None
        result = edit_cell(
            filepath=pay_file,
            sheet_name="GL Reconciliation",
            row_identifier_column="Source",
            row_identifier_value="Payroll Run Total",
            target_column="Amount",
            new_value=str(corrected_total),
            reason=f"Aligned payroll total with GL entry. {f.get('explanation', '')}",
        )
        return result.get("correction") if result.get("status") == "ok" else None

    return None


# ── tool dispatch ────────────────────────────────────────────────────────

def _handle_tool_call(name: str, args: dict) -> str:
    if name == "list_files":
        files = _all_files()
        if not files:
            return json.dumps({"files": [], "message": "No files registered. Ask the user to scan a folder."})
        summaries = []
        for fp in files:
            s = summarize_workbook(fp)
            s["path"] = str(fp)
            summaries.append(s)
        return json.dumps({"files": summaries})

    if name == "inspect_file":
        fp = _resolve(args["filename"])
        s = summarize_workbook(fp)
        s["path"] = str(fp)
        return json.dumps(s)

    if name == "load_and_audit":
        invalidate_cache()
        findings = _run_audit(args["audit_type"])
        return json.dumps({
            "audit_type": args["audit_type"],
            "total_findings": len(findings),
            "findings": findings,
        }, default=str)

    if name == "read_excel_data":
        from openpyxl import load_workbook as _load_wb
        fp = _resolve(args["filename"])
        wb = _load_wb(fp, data_only=True, read_only=True)
        sheet = args.get("sheet_name")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return json.dumps({"rows": []})
        headers = [str(h or "") for h in rows[0]]
        data_rows = [
            dict(zip(headers, [str(v) if v is not None else "" for v in r]))
            for r in rows[1:] if any(v is not None for v in r)
        ]
        return json.dumps({"sheet": ws.title, "row_count": len(data_rows), "rows": data_rows}, default=str)

    if name == "apply_correction":
        fp = _resolve(args["filename"])
        result = edit_cell(
            filepath=fp,
            sheet_name=args["sheet_name"],
            row_identifier_column=args["row_identifier_column"],
            row_identifier_value=args["row_identifier_value"],
            target_column=args["target_column"],
            new_value=args["new_value"],
            reason=args.get("reason", ""),
        )
        invalidate_cache()
        return json.dumps(result, default=str)

    if name == "auto_fix_all":
        result = _auto_fix_findings(args["audit_type"])
        return json.dumps(result, default=str)

    if name == "get_corrections_log":
        log = _writer_get_log()
        return json.dumps({"total_corrections": len(log), "corrections": log}, default=str)

    return json.dumps({"error": f"Unknown tool: {name}"})


# ── approval parsing ─────────────────────────────────────────────────────

_RE_ACTION   = re.compile(r'^APPROVAL_REQUIRED:\s*(.+)$',   re.MULTILINE)
_RE_ACCOUNTS = re.compile(r'^ACCOUNTS_AFFECTED:\s*(.+)$',   re.MULTILINE)
_RE_AMOUNT   = re.compile(r'^AMOUNT:\s*(.+)$',              re.MULTILINE)
_RE_STRIP    = re.compile(
    r'\n*^(APPROVAL_REQUIRED|ACCOUNTS_AFFECTED|AMOUNT):\s*.+$',
    re.MULTILINE,
)


def _parse_approval(text: str) -> tuple[str, dict | None]:
    action_m = _RE_ACTION.search(text)
    if not action_m:
        return text, None
    accounts_m = _RE_ACCOUNTS.search(text)
    amount_m   = _RE_AMOUNT.search(text)
    approval = {
        "action":            action_m.group(1).strip(),
        "accounts_affected": accounts_m.group(1).strip() if accounts_m else "",
        "amount":            amount_m.group(1).strip()   if amount_m   else "",
    }
    cleaned = _RE_STRIP.sub("", text).strip()
    return cleaned, approval


# ── public API ───────────────────────────────────────────────────────────

def reset():
    global conversation_history, _parsed_cache, _file_registry
    conversation_history = []
    _parsed_cache = {}
    _file_registry = {}
    _writer_clear_log()


def invalidate_cache():
    global _parsed_cache
    _parsed_cache = {}


def get_history() -> list[dict[str, str]]:
    return [m for m in conversation_history if m["role"] in ("user", "assistant")]


def get_corrections() -> list[dict]:
    return _writer_get_log()


def chat(message: str, context: dict[str, Any] | None = None) -> dict[str, Any]:
    client = _get_client()

    conversation_history.append({"role": "user", "content": message})

    system_content = SYSTEM_PROMPT
    if context:
        system_content += f"\n\n## Additional Context\n```json\n{json.dumps(context)}\n```"

    file_list = _all_files()
    if file_list:
        file_info = ", ".join(f"{fp.name} ({fp})" for fp in file_list)
        system_content += f"\n\nRegistered files: {file_info}"

    log = _writer_get_log()
    if log:
        system_content += f"\n\nCorrections already applied this session: {len(log)}"

    messages: list[dict[str, Any]] = [
        {"role": "system", "content": system_content},
        *conversation_history,
    ]

    max_rounds = 12
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    model_name = MODEL

    for _ in range(max_rounds):
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            tools=TOOLS,
            tool_choice="auto",
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
        )

        model_name = response.model
        if response.usage:
            total_usage["prompt_tokens"] += response.usage.prompt_tokens
            total_usage["completion_tokens"] += response.usage.completion_tokens
            total_usage["total_tokens"] += response.usage.total_tokens

        choice = response.choices[0]

        if choice.finish_reason == "tool_calls" or choice.message.tool_calls:
            messages.append(choice.message)

            for tc in choice.message.tool_calls:
                fn_name = tc.function.name
                fn_args = json.loads(tc.function.arguments)
                result = _handle_tool_call(fn_name, fn_args)
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc.id,
                    "content": result,
                })
            continue

        raw_reply = choice.message.content or ""
        cleaned, approval = _parse_approval(raw_reply)
        conversation_history.append({"role": "assistant", "content": cleaned})

        return {
            "reply": cleaned,
            "model": model_name,
            "usage": total_usage,
            "history_length": len(conversation_history),
            "requires_approval": approval,
        }

    fallback = "I wasn't able to complete the analysis within the allowed steps. Please try a more specific question."
    conversation_history.append({"role": "assistant", "content": fallback})
    return {
        "reply": fallback,
        "model": model_name,
        "usage": total_usage,
        "history_length": len(conversation_history),
        "requires_approval": None,
    }
