"""
Write-back utilities for correcting Excel workbooks in-place.

Every mutation is logged to a corrections ledger so there is a full
audit trail of what was changed, when, and why.
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ── corrections ledger (module-level, single-user) ───────────────────────

_corrections_log: list[dict[str, Any]] = []


def get_corrections_log() -> list[dict[str, Any]]:
    return list(_corrections_log)


def clear_corrections_log() -> None:
    global _corrections_log
    _corrections_log = []


# ── helpers ──────────────────────────────────────────────────────────────

def _find_header_row(ws: Worksheet) -> dict[str, int]:
    """Return {normalised_header: col_index} from the first row."""
    mapping: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            key = str(cell.value).strip().lower().replace(" ", "_")
            mapping[key] = col_idx
    return mapping


def _match_column(headers: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        norm = c.strip().lower().replace(" ", "_")
        if norm in headers:
            return headers[norm]
    return None


def _find_row_by_id(
    ws: Worksheet,
    id_col: int,
    id_value: str,
    max_rows: int = 500,
) -> int | None:
    """Return the 1-based row number whose id_col matches id_value."""
    for row_idx in range(2, max_rows + 2):
        cell_val = ws.cell(row=row_idx, column=id_col).value
        if cell_val is not None and str(cell_val).strip() == str(id_value).strip():
            return row_idx
    return None


# ── public API ───────────────────────────────────────────────────────────

def edit_cell(
    filepath: str | Path,
    sheet_name: str,
    row_identifier_column: str,
    row_identifier_value: str,
    target_column: str,
    new_value: Any,
    reason: str = "",
) -> dict[str, Any]:
    """
    Edit a single cell identified by (row_id_col == row_id_val, target_col).

    Returns a dict describing what was changed (old → new).
    """
    fp = Path(filepath)
    wb = load_workbook(fp)

    if sheet_name not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if candidates:
            sheet_name = candidates[0]
        else:
            wb.close()
            return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}

    ws = wb[sheet_name]
    headers = _find_header_row(ws)

    id_col = _match_column(headers, row_identifier_column)
    if id_col is None:
        wb.close()
        return {"error": f"Column '{row_identifier_column}' not found. Available: {list(headers.keys())}"}

    tgt_col = _match_column(headers, target_column)
    if tgt_col is None:
        wb.close()
        return {"error": f"Column '{target_column}' not found. Available: {list(headers.keys())}"}

    row_idx = _find_row_by_id(ws, id_col, row_identifier_value)
    if row_idx is None:
        wb.close()
        return {"error": f"No row with {row_identifier_column}='{row_identifier_value}' in sheet '{sheet_name}'"}

    cell = ws.cell(row=row_idx, column=tgt_col)
    old_value = cell.value

    if isinstance(new_value, str):
        try:
            new_value = float(new_value)
            if new_value == int(new_value):
                new_value = int(new_value)
        except ValueError:
            pass

    cell.value = new_value
    wb.save(fp)
    wb.close()

    record = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "file": fp.name,
        "sheet": sheet_name,
        "row_id": f"{row_identifier_column}={row_identifier_value}",
        "column": target_column,
        "old_value": old_value,
        "new_value": new_value,
        "reason": reason,
    }
    _corrections_log.append(record)

    return {"status": "ok", "correction": record}


def export_corrected_copy(filepath: str | Path) -> dict[str, Any]:
    """
    Create a timestamped copy of the workbook in the same directory.
    Useful for keeping the corrected version alongside the original.
    """
    fp = Path(filepath)
    if not fp.exists():
        return {"error": f"File not found: {fp.name}"}

    stem = fp.stem
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    corrected_name = f"{stem}_corrected_{ts}.xlsx"
    dest = fp.parent / corrected_name

    shutil.copy2(fp, dest)
    return {"status": "ok", "corrected_file": corrected_name, "path": str(dest)}
